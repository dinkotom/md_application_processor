
import unittest
import sys
import os
import sqlite3
import openpyxl
import json
from io import BytesIO

# Add project root to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from web_app import app
from src.database import get_db_connection

class TestExportStatus(unittest.TestCase):
    
    def setUp(self):
        # Configure app for testing
        app.config['TESTING'] = True
        app.config['SECRET_KEY'] = 'test-key'
        self.client = app.test_client()
        
        # Setup temporary test database
        self.db_path = 'test_export_status.db'
        if os.path.exists(self.db_path):
            os.remove(self.db_path)
        
        # Initialize database schema
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Schema for applicants
        cursor.execute('''
            CREATE TABLE applicants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                first_name TEXT,
                last_name TEXT,
                email TEXT,
                status TEXT,
                deleted INTEGER DEFAULT 0,
                newsletter INTEGER DEFAULT 0,
                email_sent INTEGER DEFAULT 0,
                phone_warning_dismissed INTEGER DEFAULT 0,
                parent_email_warning_dismissed INTEGER DEFAULT 0,
                duplicate_warning_dismissed INTEGER DEFAULT 0,
                exported_to_ecomail INTEGER DEFAULT 0
            )
        ''')
        
        # Schema for export_presets with filter_status
        cursor.execute('''
            CREATE TABLE export_presets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                fields TEXT NOT NULL,
                filter_status TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Insert test data with different statuses
        applicants_data = [
            ('Jan', 'Novák', 'jan@test.cz', 'Nová'),
            ('Petr', 'Svoboda', 'petr@test.cz', 'Vyřízená'),
            ('Marie', 'Dvořáková', 'marie@test.cz', 'Zpracovává se'),
            ('Eva', 'Kovářová', 'eva@test.cz', 'Nová')
        ]
        
        cursor.executemany('''
            INSERT INTO applicants (first_name, last_name, email, status)
            VALUES (?, ?, ?, ?)
        ''', applicants_data)
        
        conn.commit()
        conn.close()
        
        # Patch get_db_path using unittest.mock
        from unittest.mock import patch
        self.db_patcher = patch('src.database.get_db_path', return_value=self.db_path)
        self.db_patcher.start()
        
        # Simulate logged-in user
        with self.client.session_transaction() as sess:
            sess['user'] = {'email': 'tester@example.com', 'name': 'Tester'}

    def tearDown(self):
        self.db_patcher.stop()
        if os.path.exists(self.db_path):
           os.remove(self.db_path)

    def test_filter_by_single_status(self):
        """Test export filtered by single status"""
        response = self.client.post('/export/excel', data={
            'fields': ['first_name', 'status'],
            'status': 'Vyřízená'
        })
        self.assertEqual(response.status_code, 200)
        
        wb = openpyxl.load_workbook(BytesIO(response.data))
        ws = wb.active
        
        # Header is row 1. Data starts row 2.
        # Should contain only Petr Svoboda (Vyřízená)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], 'Petr')
        self.assertEqual(rows[0][1], 'Vyřízená')

    def test_filter_by_multiple_statuses(self):
        """Test export filtered by multiple statuses"""
        # Simulate multi-value form submission for status
        data = [
            ('fields', 'first_name'),
            ('fields', 'status'),
            ('status', 'Nová'),
            ('status', 'Zpracovává se')
        ]
        # In werkzeug test client, passing a list of tuples or dict with list values works for MultiDict
        # Using dict with list for test client:
        response = self.client.post('/export/excel', data={
            'fields': ['first_name', 'status'],
            'status': ['Nová', 'Zpracovává se']
        })
        self.assertEqual(response.status_code, 200)
        
        wb = openpyxl.load_workbook(BytesIO(response.data))
        ws = wb.active
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Should include Jan (Nová), Marie (Zpracovává se), Eva (Nová). Total 3.
        # Petr (Vyřízená) should be excluded.
        self.assertEqual(len(rows), 3)
        statuses = [r[1] for r in rows]
        self.assertIn('Nová', statuses)
        self.assertIn('Zpracovává se', statuses)
        self.assertNotIn('Vyřízená', statuses)

    def test_preset_status_persistence(self):
        """Test that status filter is saved and retrieved in presets"""
        preset_data = {
            'name': 'Status Filter Preset',
            'fields': ['first_name', 'email'],
            'status_filter': ['Nová', 'Vyřízená']
        }
        
        # Save
        response = self.client.post('/export/presets', json=preset_data)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json['success'])
        
        # List
        response = self.client.get('/export/presets')
        self.assertEqual(response.status_code, 200)
        presets = response.json['presets']
        self.assertEqual(len(presets), 1)
        
        saved_preset = presets[0]
        self.assertEqual(saved_preset['name'], 'Status Filter Preset')
        
        # Check filter_status (it comes back as a JSON string from DB via dict(row))
        # Wait, my route implementation returns `dict(p)` where p is sqlite3.Row.
        # So filter_status will be the string stored in DB.
        self.assertIsNotNone(saved_preset.get('filter_status'))
        
        loaded_status = json.loads(saved_preset['filter_status'])
        self.assertEqual(loaded_status, ['Nová', 'Vyřízená'])

if __name__ == '__main__':
    unittest.main()
