import unittest
import sys
import os
import sqlite3
import openpyxl
from io import BytesIO
from datetime import datetime, date

# Add project root to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from web_app import app
from src.database import get_db_connection

class TestExcelExport(unittest.TestCase):
    
    def setUp(self):
        # Configure app for testing
        app.config['TESTING'] = True
        app.config['SECRET_KEY'] = 'test-key'
        self.client = app.test_client()
        
        # Setup temporary test database
        self.db_path = 'test_excel_export.db'
        if os.path.exists(self.db_path):
            os.remove(self.db_path)
        
        # Initialize database schema
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Minimal schema for applicants
        cursor.execute('''
            CREATE TABLE applicants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                first_name TEXT,
                last_name TEXT,
                email TEXT,
                phone TEXT,
                dob TEXT,
                membership_id TEXT,
                city TEXT,
                school TEXT,
                status TEXT,
                application_received TEXT,
                created_at TEXT,
                interests TEXT,
                character TEXT,
                frequency TEXT,
                color TEXT,
                source TEXT,
                source_detail TEXT,
                message TEXT,
                newsletter INTEGER DEFAULT 0,
                guessed_gender TEXT,
                email_sent INTEGER DEFAULT 0,
                email_sent_at TEXT,
                note TEXT,
                deleted INTEGER DEFAULT 0,
                phone_warning_dismissed INTEGER DEFAULT 0
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE export_presets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                fields TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Insert test data
        cursor.execute('''
            INSERT INTO applicants (
                first_name, last_name, email, phone, dob, 
                membership_id, city, school, status, 
                application_received, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            'Test', 'User', 'test@example.com', '+420123456789', '01.01.2000',
            '1001', 'Prague', 'Test School', 'Nová',
            '2025-12-16 12:00:00', '2025-12-16 11:00:00'
        ))
        
        conn.commit()
        conn.close()
        
        # Patch get_db_path using unittest.mock
        from unittest.mock import patch
        self.db_patcher = patch('src.database.get_db_path', return_value=self.db_path)
        self.db_patcher.start()
        
        # Simulate logged-in user
        with self.client.session_transaction() as sess:
            sess['user'] = {'email': 'test@example.com', 'name': 'Test User'}

    def tearDown(self):
        self.db_patcher.stop()
        if os.path.exists(self.db_path):
            os.remove(self.db_path)

    def test_export_no_fields(self):
        """Test error when no fields are selected"""
        response = self.client.post('/export/excel', data={})
        self.assertEqual(response.status_code, 400)
        self.assertIn(b'Chyba', response.data)

    def test_export_fields_order(self):
        """Test that exported columns match selected fields and order"""
        # Select specific fields in specific order
        fields = ['first_name', 'email', 'membership_id']
        response = self.client.post('/export/excel', data={
            'fields': fields
        })
        
        self.assertEqual(response.status_code, 200)
        
        # Load Excel
        wb = openpyxl.load_workbook(BytesIO(response.data))
        ws = wb.active
        
        # Verify Headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = ['Jméno', 'Email', 'Členské číslo']
        self.assertEqual(headers, expected_headers)
        
        # Verify Data
        row = [cell.value for cell in ws[2]]
        self.assertEqual(row[0], 'Test')
        self.assertEqual(row[1], 'test@example.com')
        self.assertEqual(row[2], '1001')

    def test_export_date_formatting(self):
        """Test that dates are exported as objects and formatted"""
        fields = ['dob', 'application_received', 'created_at']
        response = self.client.post('/export/excel', data={
            'fields': fields
        })
        
        wb = openpyxl.load_workbook(BytesIO(response.data))
        ws = wb.active
        
        # Check DOB (Date object)
        dob_cell = ws.cell(row=2, column=1)
        self.assertIsInstance(dob_cell.value, (datetime, date))
        self.assertEqual(dob_cell.number_format, 'DD.MM.YYYY')
        self.assertEqual(dob_cell.value.strftime('%Y-%m-%d'), '2000-01-01')
        
        # Check Application Received (Datetime object)
        received_cell = ws.cell(row=2, column=2)
        self.assertIsInstance(received_cell.value, datetime)
        self.assertEqual(received_cell.number_format, 'DD.MM.YYYY HH:MM:SS')
        self.assertEqual(received_cell.value.strftime('%Y-%m-%d %H:%M:%S'), '2025-12-16 12:00:00')

        # Check Application Received (Datetime object)
        received_cell = ws.cell(row=2, column=2)
        self.assertIsInstance(received_cell.value, datetime)
        self.assertEqual(received_cell.number_format, 'DD.MM.YYYY HH:MM:SS')
        self.assertEqual(received_cell.value.strftime('%Y-%m-%d %H:%M:%S'), '2025-12-16 12:00:00')

    def test_export_presets(self):
        """Test creating, listing and deleting export presets"""
        
        # 1. Create Preset
        import json
        preset_data = {
            'name': 'My Special Export',
            'fields': ['email', 'phone', 'dob']
        }
        
        # Need to ensure the table was created in the test DB inside setUp if not migrated?
        # Our migration logic is separate. Ideally setUp should initialize schema.
        # Let's add the table creation to setUp of this test file to be safe self-contained.
        # Update: setUp already has table creation for applicants, we need to add export_presets.
        
        response = self.client.post('/export/presets', json=preset_data)
        
        # If table doesn't exist, this will fail 500.
        # We need to updating setUp first.
        
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json['success'])
        
        # 2. List Presets
        response = self.client.get('/export/presets')
        self.assertEqual(response.status_code, 200)
        data = response.json
        self.assertEqual(len(data['presets']), 1)
        self.assertEqual(data['presets'][0]['name'], 'My Special Export')
        fields_loaded = json.loads(data['presets'][0]['fields'])
        self.assertEqual(fields_loaded, ['email', 'phone', 'dob'])
        
        preset_id = data['presets'][0]['id']
        
        # 3. Delete Preset
        response = self.client.delete(f'/export/presets/{preset_id}')
        self.assertEqual(response.status_code, 200)
        
        # Verify Empty
        response = self.client.get('/export/presets')
        self.assertEqual(len(response.json['presets']), 0)

if __name__ == '__main__':
    unittest.main()
