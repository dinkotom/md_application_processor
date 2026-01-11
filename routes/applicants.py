from flask import Blueprint, render_template, request, session, redirect, url_for, jsonify, send_file, make_response
from src.database import get_db_connection, log_action, get_db_path
from src.validator import is_valid_email, is_valid_phone, is_suspect_parent_email, check_duplicate_contact
from src.parser import normalize_phone, normalize_school, calculate_age
from src.generator import generate_card
from datetime import datetime, date
from io import BytesIO
import json
import logging

# Define Blueprint
applicants_bp = Blueprint('applicants', __name__)
logger = logging.getLogger(__name__)

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

def get_filtered_applicants(request_args):
    """Helper to get filtered and sorted applicants based on request args"""
    conn = get_db_connection()
    
    # Get parameters
    search = request_args.get('search', '')
    # Handle status as list if multiple values provided (e.g. from export form checkboxes)
    filter_status = request_args.getlist('status') if hasattr(request_args, 'getlist') else [request_args.get('status')]
    # Clean up empty strings and single-item list that is effectively empty
    filter_status = [s for s in filter_status if s]
    
    filter_age_group = request_args.get('age_group', '')
    filter_city = request_args.get('city', '')
    filter_school = request_args.get('school', '')
    filter_interest = request_args.get('interest', '')
    filter_source = request_args.get('source', '')
    filter_alerts = request_args.get('alerts', '')
    filter_character = request_args.get('character', '')
    filter_guessed_gender = request_args.get('guessed_gender', '')
    sort_by = request_args.get('sort', 'id')
    sort_order = request_args.get('order', 'desc')
    
    # Build query
    query = "SELECT * FROM applicants WHERE deleted = 0 AND 1=1"
    params = []
    
    if search:
        # Use remove_diacritics for diacritic-insensitive search
        query += " AND (remove_diacritics(first_name) LIKE remove_diacritics(?) OR remove_diacritics(last_name) LIKE remove_diacritics(?) OR remove_diacritics(email) LIKE remove_diacritics(?) OR remove_diacritics(city) LIKE remove_diacritics(?))"
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param, search_param])
    
    if filter_status:
        # Check if it's a list or single value
        if isinstance(filter_status, list) and len(filter_status) > 0:
            placeholders = ','.join(['?'] * len(filter_status))
            query += f" AND status IN ({placeholders})"
            params.extend(filter_status)
        elif isinstance(filter_status, str):
            query += " AND status = ?"
            params.append(filter_status)

    if filter_source:
        query += " AND source = ?"
        params.append(filter_source)

    if filter_character:
        query += " AND character = ?"
        params.append(filter_character)

    if filter_guessed_gender:
        query += " AND guessed_gender = ?"
        params.append(filter_guessed_gender)

    if filter_interest:
        query += " AND interests LIKE ?"
        params.append(f"%{filter_interest}%")

    cursor = conn.execute(query, params)
    all_applicants = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # Python-side filtering
    if filter_city:
        all_applicants = [app for app in all_applicants if app.get('city') and app['city'].strip().lower() == filter_city.lower()]
    
    if filter_school:
        all_applicants = [app for app in all_applicants if app.get('school') and normalize_school(app['school']) == filter_school]
    
    # Filter by age
    if filter_age_group:
        filtered = []
        for app in all_applicants:
            age = calculate_age(app['dob']) if app.get('dob') else None
            if age is not None:
                if filter_age_group == 'under_15' and age < 15: filtered.append(app)
                elif filter_age_group == '15_18' and 15 <= age <= 18: filtered.append(app)
                elif filter_age_group == '19_24' and 19 <= age <= 24: filtered.append(app)
                elif filter_age_group == 'over_24' and age > 24: filtered.append(app)
        all_applicants = filtered

    # Filter by alerts
    if filter_alerts == 'true':
        with_alerts = []
        for app in all_applicants:
            has_alert = False
            age = calculate_age(app['dob']) if app.get('dob') else None
            if age is None or age < 15 or age >= 25:
                has_alert = True
            
            if not app.get('parent_email_warning_dismissed') and is_suspect_parent_email(app.get('first_name',''), app.get('last_name',''), app.get('email','')):
                has_alert = True
                
            if not app.get('duplicate_warning_dismissed'):
                 # Check duplicates (this might be slow for many records but we are filtering already)
                 # Wait, 'is_duplicate' is usually calculated later in the view.
                 # Recomputing here might be expensive. But we need it for filtering.
                 # Let's verify duplication logic.
                 duplicates = check_duplicate_contact(
                    app.get('email', ''),
                    app.get('phone', ''),
                    current_id=app.get('id'),
                    db_path=get_db_path()
                )
                 if duplicates['email_duplicate'] or duplicates['phone_duplicate']:
                     has_alert = True
            
            # Check validity
            if not is_valid_email(app.get('email', '')): has_alert = True
            if not is_valid_phone(app.get('phone', '')) and not app.get('phone_warning_dismissed'): has_alert = True
            
            if has_alert:
                with_alerts.append(app)
        all_applicants = with_alerts

    # Sort
    if sort_by == 'application_received':
        all_applicants.sort(
            key=lambda x: (x.get('application_received') is None, x.get('application_received') or ''),
            reverse=(sort_order == 'desc')
        )
    else:  # Default to ID (membership_id)
        all_applicants.sort(
            key=lambda x: (
                0 if str(x.get('membership_id', '')).isdigit() else 1, 
                int(x.get('membership_id')) if str(x.get('membership_id', '')).isdigit() else x.get('membership_id', '')
            ), 
            reverse=(sort_order == 'desc')
        )
    return all_applicants

# --- Routes ---

@applicants_bp.route('/')
@login_required
def index():
    """Main dashboard page"""
    search = request.args.get('search', '')
    filter_status = request.args.get('status', '')
    filter_age_group = request.args.get('age_group', '')
    filter_city = request.args.get('city', '')
    filter_school = request.args.get('school', '')
    filter_interest = request.args.get('interest', '')
    filter_source = request.args.get('source', '')
    filter_alerts = request.args.get('alerts', '')
    filter_character = request.args.get('character', '')
    filter_guessed_gender = request.args.get('guessed_gender', '')
    
    all_applicants = get_filtered_applicants(request.args)
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20
    total_pages = (len(all_applicants) + per_page - 1) // per_page
    start = (page - 1) * per_page
    end = start + per_page
    applicants_subset = all_applicants[start:end]
    
    final_applicants = []
    for app in applicants_subset:
        app['age'] = calculate_age(app['dob']) if app.get('dob') else None
        
        # Check alerts
        app['suspect_parent_email'] = (
            is_suspect_parent_email(
                app.get('first_name', ''),
                app.get('last_name', ''),
                app.get('email', '')
            ) and not app.get('parent_email_warning_dismissed', 0)
        )
        
        duplicates = check_duplicate_contact(
            app.get('email', ''),
            app.get('phone', ''),
            current_id=app.get('id'),
            db_path=get_db_path()
        )
        app['is_duplicate'] = (
            (duplicates['email_duplicate'] or duplicates['phone_duplicate']) 
            and not app.get('duplicate_warning_dismissed', 0)
        )
        
        app['invalid_email'] = not is_valid_email(app.get('email', ''))
        app['invalid_phone'] = not is_valid_phone(app.get('phone', '')) and not app.get('phone_warning_dismissed', 0)

        final_applicants.append(app)
    
    return render_template('index.html', 
                         applicants=final_applicants, 
                         search=search,
                         filter_status=filter_status,
                         filter_age_group=filter_age_group,
                         filter_city=filter_city,
                         filter_school=filter_school,
                         filter_interest=filter_interest,
                         filter_source=filter_source,
                         filter_alerts=filter_alerts,
                         filter_character=filter_character,
                         page=page,
                         total_pages=total_pages,
                         total=len(all_applicants))

@applicants_bp.route('/applicant/<int:id>')
@login_required
def detail(id):
    """Applicant detail page"""
    conn = get_db_connection()
    applicant = conn.execute('SELECT * FROM applicants WHERE id = ?', (id,)).fetchone()
    
    audit_logs = conn.execute('SELECT * FROM audit_logs WHERE applicant_id = ? ORDER BY timestamp DESC', (id,)).fetchall()
    
    prev_id = conn.execute('SELECT id FROM applicants WHERE id < ? AND deleted = 0 ORDER BY id DESC LIMIT 1', (id,)).fetchone()
    next_id = conn.execute('SELECT id FROM applicants WHERE id > ? AND deleted = 0 ORDER BY id ASC LIMIT 1', (id,)).fetchone()
    conn.close()
    
    if applicant is None:
        return 'Applicant not found', 404
        
    app_dict = dict(applicant)
    app_dict['age'] = calculate_age(app_dict.get('dob', '')) if app_dict.get('dob') else None
    
    # Check alerts for detail view
    app_dict['suspect_parent_email'] = (
        is_suspect_parent_email(
            app_dict.get('first_name', ''),
            app_dict.get('last_name', ''),
            app_dict.get('email', '')
        ) and not app_dict.get('parent_email_warning_dismissed', 0)
    )
    
    duplicates = check_duplicate_contact(
        app_dict.get('email', ''),
        app_dict.get('phone', ''),
        current_id=app_dict.get('id'),
        db_path=get_db_path()
    )
    app_dict['is_duplicate'] = (
        (duplicates['email_duplicate'] or duplicates['phone_duplicate']) 
        and not app_dict.get('duplicate_warning_dismissed', 0)
    )
    app_dict['duplicate_details'] = duplicates
    
    app_dict['invalid_email'] = not is_valid_email(app_dict.get('email', ''))
    app_dict['invalid_phone'] = not is_valid_phone(app_dict.get('phone', '')) and not app_dict.get('phone_warning_dismissed', 0)
    
    # Determine Ecomail list name for confirmation modal
    mode = session.get('mode', 'test')
    ecomail_list_id = 17 if session.get('mode') == 'test' else 16
    ecomail_list_name = f"{ecomail_list_id} ({session.get('mode', 'test').capitalize()})"
    
    return render_template('detail.html', 
                           applicant=app_dict, 
                           audit_logs=audit_logs,
                           ecomail_list_name=ecomail_list_name,
                           ecomail_list_id=ecomail_list_id,
                           prev_id=prev_id['id'] if prev_id else None,
                           next_id=next_id['id'] if next_id else None,
                           back_args=request.args)

@applicants_bp.route('/applicant/<int:id>/update_field', methods=['POST'])
@login_required
def update_applicant_field(id):
    """Update a specific field via AJAX"""
    data = request.json
    field = data.get('field')
    value = data.get('value')
    
    # Map 'name' to 'first_name' for compatibility
    if field == 'name':
        field = 'first_name'
        
    allowed_fields = [
        'first_name', 'last_name', 'email', 'phone', 'dob', 'city', 
        'school', 'interests', 'character', 'status', 'newsletter',
        'source', 'source_detail', 'message', 'note', 'guessed_gender',
        'parent_email_warning_dismissed', 'duplicate_warning_dismissed', 'phone_warning_dismissed'
    ]
    
    if field not in allowed_fields:
        return jsonify({'success': False, 'error': 'Invalid field'}), 400
        
    conn = get_db_connection()
    try:
        # Get old value for audit log
        old_row = conn.execute(f'SELECT {field} FROM applicants WHERE id = ?', (id,)).fetchone()
        old_value = old_row[0] if old_row else None
        
        # Special validations
        if field == 'email' and not is_valid_email(value):
             return jsonify({'success': False, 'error': 'Invalid email format'}), 400
        
        if field == 'phone':
            value = normalize_phone(value)
            if not is_valid_phone(value):
                return jsonify({'success': False, 'error': 'Invalid phone format (min 9 digits)'}), 400
             
        conn.execute(f'UPDATE applicants SET {field} = ? WHERE id = ?', (value, id))
        conn.commit()
        
        log_action(id, f"Uprava pole {field}", session['user']['email'], old_value, value)
        
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error updating applicant {id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@applicants_bp.route('/applicant/<int:id>/delete', methods=['POST'])
@login_required
def delete_applicant(id):
    """Soft delete an applicant"""
    conn = get_db_connection()
    conn.execute('UPDATE applicants SET deleted = 1 WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    if 'user' in session:
        log_action(id, "Smazání přihlášky", session['user']['email'])
        
    next_url = request.form.get('next') or request.args.get('next')
    return redirect(next_url or url_for('applicants.index'))

@applicants_bp.route('/applicant/<int:id>/status', methods=['POST'])
@login_required
def update_applicant_status(id):
    """Update applicant status"""
    new_status = request.form.get('status')
    if new_status:
        conn = get_db_connection()
        old_status = conn.execute('SELECT status FROM applicants WHERE id = ?', (id,)).fetchone()['status']
        conn.execute('UPDATE applicants SET status = ? WHERE id = ?', (new_status, id))
        conn.commit()
        conn.close()
        
        log_action(id, "Změna stavu", session['user']['email'], old_status, new_status)
        
    next_url = request.form.get('next') or request.args.get('next')
    return redirect(next_url or request.referrer or url_for('applicants.index'))

@applicants_bp.route('/applicant/<int:id>/dismiss-parent-warning', methods=['POST'])
@login_required
def dismiss_parent_warning(id):
    """Dismiss parent email warning"""
    conn = get_db_connection()
    conn.execute('UPDATE applicants SET parent_email_warning_dismissed = 1 WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('applicants.detail', id=id))

@applicants_bp.route('/applicant/<int:id>/dismiss-duplicate-warning', methods=['POST'])
@login_required
def dismiss_duplicate_warning(id):
    """Dismiss duplicate contact warning"""
    conn = get_db_connection()
    conn.execute('UPDATE applicants SET duplicate_warning_dismissed = 1 WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('applicants.detail', id=id))

@applicants_bp.route('/applicant/<int:id>/dismiss-phone-warning', methods=['POST'])
@login_required
def dismiss_phone_warning(id):
    """Dismiss invalid phone format warning"""
    conn = get_db_connection()
    conn.execute('UPDATE applicants SET phone_warning_dismissed = 1 WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('applicants.detail', id=id))

@applicants_bp.route('/applicant/<int:id>/card')
@login_required
def applicant_card(id):
    """Generate membership card"""
    conn = get_db_connection()
    applicant = conn.execute('SELECT * FROM applicants WHERE id = ?', (id,)).fetchone()
    conn.close()
    
    if not applicant:
        return "Applicant not found", 404
        
    try:
        img_io = generate_card(dict(applicant))
        return send_file(img_io, mimetype='image/png', as_attachment=False, download_name=f'card_{id}.png')
    except Exception as e:
        logger.error(f"Error generating card for applicant {id}: {e}")
        return "Error generating card", 500

@applicants_bp.route('/applicant/<int:id>/card_preview')
@login_required
def serve_card_preview(id):
    """Serve card for preview image"""
    return applicant_card(id)

@applicants_bp.route('/applicant/<int:id>/send_welcome_email', methods=['POST'])
@login_required
def send_welcome_email_route(id):
    """Send welcome email with card"""
    import os
    from src.generator import generate_card
    from src.email_sender import send_welcome_email
    
    conn = get_db_connection()
    applicant = conn.execute('SELECT * FROM applicants WHERE id = ?', (id,)).fetchone()
    
    if not applicant:
        conn.close()
        return jsonify({'success': False, 'error': 'Applicant not found'}), 404
        
    app_data = dict(applicant)
    
    # Generate card bytes
    try:
        card_bytes = generate_card(app_data)
    except Exception as e:
        conn.close()
        logger.error(f"Error generating card for email: {e}")
        return jsonify({'success': False, 'error': f'Card generation failed: {str(e)}'}), 500

    # Get credentials
    email_user = os.getenv('SMTP_USER')
    email_pass = os.getenv('SMTP_PASS')
    smtp_host = os.getenv('SMTP_HOST', 'smtp.gmail.com')
    smtp_port = os.getenv('SMTP_PORT', 465)
    
    if not email_user or not email_pass:
        conn.close()
        return jsonify({'success': False, 'error': 'Email credentials not configured'}), 500
        
    # Determine mode from session
    mode = session.get('mode', 'test')
    
    # In test mode, we want to copy the logged-in user
    user_email = session.get('user', {}).get('email')
    
    result = send_welcome_email(
        app_data, 
        card_bytes, 
        email_user, 
        email_pass, 
        mode=mode, 
        copy_to=user_email,
        smtp_host=smtp_host,
        smtp_port=smtp_port
    )
    
    if result['success']:
        # Update DB
        conn.execute('UPDATE applicants SET email_sent = 1, email_sent_at = ? WHERE id = ?', 
                    (datetime.now(), id))
        conn.commit()
        log_action(id, "Odeslán uvítací email", session['user']['email'])
    
    conn.close()
    
    return jsonify({
        'success': result['success'],
        'mode': mode,
        'recipient': result.get('recipient'),
        'error': result.get('error'),
        'message': result.get('message')
    })

def _prepare_ecomail_data(app_data):
    """Helper to prepare subscriber data from applicant dict"""
    email = app_data.get('email')
    
    # Prepare tags
    tags = []
    
    # Character as tag
    char_val = app_data.get('character')
    if char_val:
        for char_item in str(char_val).split(','):
            cleaned = char_item.strip()
            if cleaned:
                tags.append(cleaned)
        
    # Interests as tags (comma separated)
    interests_val = app_data.get('interests')
    if interests_val:
        for interest in interests_val.split(','):
            cleaned = interest.strip()
            if cleaned:
                tags.append(cleaned)

    # School as tag
    school_val = app_data.get('school')
    if school_val:
        tags.append(str(school_val).replace(',', ' '))

    # Prepare subscriber data
    subscriber_data = {
        'email': email,
        'name': app_data.get('first_name', ''),
        'surname': app_data.get('last_name', ''),
        'phone': app_data.get('phone', ''),
        'city': app_data.get('city', ''),
        'tags': tags,
        # Custom fields
        'custom_fields': {
            'MEMBERSHIP_ID': app_data.get('membership_id', '')
        }
    }
    
    # Handle Birthday (Native field, format YYYY-MM-DD)
    dob = app_data.get('dob')
    if dob:
        try:
            # Assuming input format DD.MM.YYYY or DD/MM/YYYY
            dob_clean = dob.strip().replace('/', '.')
            dt = datetime.strptime(dob_clean, '%d.%m.%Y')
            subscriber_data['birthday'] = dt.strftime('%Y-%m-%d')
        except Exception:
            # If parsing fails, try sending raw string or log warning
            subscriber_data['birthday'] = dob
            
    return subscriber_data

@applicants_bp.route('/applicant/<int:id>/check_ecomail', methods=['GET'])
@login_required
def check_ecomail(id):
    """Check status of applicant in Ecomail before export"""
    from src.ecomail import EcomailClient
    
    conn = get_db_connection()
    applicant = conn.execute('SELECT * FROM applicants WHERE id = ?', (id,)).fetchone()
    conn.close()
    
    if not applicant:
        return jsonify({'success': False, 'error': 'Applicant not found'}), 404
        
    app_data = dict(applicant)
    email = app_data.get('email')
    
    if not email:
        return jsonify({'success': False, 'error': 'No email address'}), 400
        
    mode = session.get('mode', 'test')
    
    try:
        client = EcomailClient()
        # Check if subscriber exists
        existing_response = client.get_subscriber(email)
        existing_data = existing_response.get('data', {}).get('subscriber', {}) if existing_response.get('success') else None
        
        # Prepare proposed data
        proposed_data = _prepare_ecomail_data(app_data)
        
        # Perform Comparison
        has_changes = False
        diff = []
        
        if existing_data:
            # Helper to normalize values for comparison
            def normalize(val):
                if val is None: return ""
                return str(val).strip()

            # Fields to compare
            # Label, Key in Existing, Key in Proposed, Type
            fields_map = [
                ('Jméno', 'name', 'name'),
                ('Příjmení', 'surname', 'surname'),
                ('Email', 'email', 'email', 'email'),
                ('Telefon', 'phone', 'phone', 'phone'),
                ('Datum narození', 'birthday', 'birthday'),
                ('Štítky (Tags)', 'tags', 'tags', 'tags'),
            ]
            
            # 1. Compare standard fields
            for label, e_key, p_key, *fmt in fields_map:
                e_val = existing_data.get(e_key)
                p_val = proposed_data.get(p_key)
                field_type = fmt[0] if fmt else 'string'
                
                is_diff = False
                
                if field_type == 'tags':
                    # Tags are lists
                    e_tags = set(t.strip() for t in (e_val or []))
                    p_tags = set(t.strip() for t in (p_val or []))
                    if e_tags != p_tags:
                        is_diff = True
                    e_display = ", ".join(sorted(e_tags)) or "-"
                    p_display = ", ".join(sorted(p_tags)) or "-"
                elif field_type == 'email':
                   if normalize(e_val).lower() != normalize(p_val).lower():
                       is_diff = True
                   e_display = e_val or "-"
                   p_display = p_val or "-"
                elif field_type == 'phone':
                   # Strip spaces
                   if normalize(e_val).replace(" ", "") != normalize(p_val).replace(" ", ""):
                       is_diff = True
                   e_display = e_val or "-"
                   p_display = p_val or "-"
                else:
                    # Generic string
                    if normalize(e_val) != normalize(p_val):
                        is_diff = True
                    e_display = e_val or "-"
                    p_display = p_val or "-"
                
                if is_diff:
                    has_changes = True
                    
                diff.append({
                    'label': label,
                    'existing': e_display,
                    'proposed': p_display,
                    'is_diff': is_diff
                })
            
            # 2. Compare Custom Fields (MEMBERSHIP_ID)
            # Find MEMBERSHIP_ID in existing custom fields
            # It might be in 'custom_fields' dict or 'lists' array/dict depending on structure
            
            # Helper to get custom fields safely
            e_custom = existing_data.get('custom_fields') or {}
            
            # If empty, try to find in lists (Ecomail API v2 often returns this structure)
            if not e_custom:
                lists = existing_data.get('lists')
                if lists and isinstance(lists, dict):
                    # Lists is keyed by ID string
                    target_list_id = '16' if mode == 'production' else '17'
                    list_data = lists.get(target_list_id)
                    if list_data:
                         e_custom = list_data.get('c_fields') or {}
                elif lists and isinstance(lists, list):
                     # Fallback if lists is array
                     target_list_id = 16 if mode == 'production' else 17
                     for l in lists:
                         if l.get('list_id') == target_list_id:
                             e_custom = l.get('custom_fields') or l.get('c_fields') or {}
                             break

            e_mem_id = e_custom.get('MEMBERSHIP_ID') # Key MUST match what is in Ecomail (case sensitive?)
            # Fallback for old key if needed
            if not e_mem_id:
                 # Check for 0 as well? Sometimes it is 0.
                 # debug script showed CLENSKE_CISLO: 0. MEMBERSHIP_ID: 2022.
                 e_mem_id = e_custom.get('clenske_cislo') or e_custom.get('CLENSKE_CISLO')

            p_mem_id = proposed_data.get('custom_fields', {}).get('MEMBERSHIP_ID')
            
            is_mem_diff = normalize(e_mem_id) != normalize(p_mem_id)
            if is_mem_diff:
                has_changes = True
                
            diff.append({
                'label': 'Členské číslo',
                'existing': e_mem_id or "-",
                'proposed': p_mem_id or "-",
                'is_diff': is_mem_diff
            })

        return jsonify({
            'success': True,
            'exists': existing_response.get('success', False),
            'existing_data': existing_data,
            'proposed_data': proposed_data,
            'has_changes': has_changes,
            'diff': diff,
            'mode': mode,
            'list_id': '16' if mode == 'production' else '17'
        })
        
    except Exception as e:
        logger.error(f"Ecomail check error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@applicants_bp.route('/applicant/<int:id>/export_to_ecomail', methods=['POST'])
@login_required
def export_to_ecomail(id):
    """Export applicant to Ecomail"""
    import os
    from src.ecomail import EcomailClient
    
    conn = get_db_connection()
    applicant = conn.execute('SELECT * FROM applicants WHERE id = ?', (id,)).fetchone()
    
    if not applicant:
        conn.close()
        return jsonify({'success': False, 'error': 'Applicant not found'}), 404
        
    app_data = dict(applicant)
    email = app_data.get('email')
    
    if not email:
        conn.close()
        return jsonify({'success': False, 'error': 'No email address'}), 400
        
    mode = session.get('mode', 'test')
    
    # Select List ID based on mode
    if mode == 'production':
        list_id = '16'
    else:
        list_id = '17'

    # list_id = os.getenv('ECOMAIL_LIST_ID') # Deprecated in favor of hardcoded mode logic per request
        
    try:
        client = EcomailClient()
        
        # Prepare subscriber data using helper
        subscriber_data = _prepare_ecomail_data(app_data)
        
        # Determine newsletter status (1=subscribed, 0=unsubscribed)
        # We pass this to create_subscriber, which decides whether to use it (only for new users)
        newsletter_consent = app_data.get('newsletter', 1) 
        
        result = client.create_subscriber(list_id, subscriber_data, newsletter_status=newsletter_consent)
        
        if result['success']:
            conn.execute('UPDATE applicants SET exported_to_ecomail = 1, exported_at = ? WHERE id = ?', 
                        (datetime.now(), id))
            conn.commit()
            log_action(id, "Export do Ecomailu", session['user']['email'])
            
        conn.close()
        return jsonify(result)
        
    except Exception as e:
        conn.close()
        logger.error(f"Ecomail export error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Export Logic - Simple Excel Export
@applicants_bp.route('/exports')
@login_required
def exports():
    """Exports page"""
    return render_template('exports.html')

# -- Export Presets API --

@applicants_bp.route('/export/presets', methods=['GET'])
@login_required
def get_export_presets():
    """List all saved export presets"""
    conn = get_db_connection()
    presets = conn.execute('SELECT * FROM export_presets ORDER BY name').fetchall()
    conn.close()
    
    return jsonify({
        'success': True,
        'presets': [dict(p) for p in presets]
    })

@applicants_bp.route('/export/presets', methods=['POST'])
@login_required
def save_export_preset():
    """Save a new export preset"""
    data = request.get_json()
    name = data.get('name')
    fields = data.get('fields') # List of field IDs
    status_filter = data.get('status_filter') # List of statuses (optional)
    
    if not name or not fields:
        return jsonify({'success': False, 'error': 'Missing name or fields'}), 400
        
    import json
    fields_json = json.dumps(fields)
    status_json = json.dumps(status_filter) if status_filter else None
    
    conn = get_db_connection()
    try:
        conn.execute('INSERT INTO export_presets (name, fields, filter_status) VALUES (?, ?, ?)', (name, fields_json, status_json))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500
        
    conn.close()
    return jsonify({'success': True})

@applicants_bp.route('/export/presets/<int:id>', methods=['DELETE'])
@login_required
def delete_export_preset(id):
    """Delete an export preset"""
    conn = get_db_connection()
    conn.execute('DELETE FROM export_presets WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# -- Excel Export --
@applicants_bp.route('/export/excel', methods=['POST'])
@login_required
def export_excel():
    """Export filtered applicants to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    # Get selected fields from form
    selected_fields = request.form.getlist('fields')
    # Status is already handled by get_filtered_applicants reading from request.form/request.args
    # Note: request.args is usually GET, request.form is POST. 
    # get_filtered_applicants takes "request_args" which can be a CombinedMultiDict or just request.values
    # We should pass request.values to cover both or construct a dict
    
    # We need to construct args that include the POST form data for status checkboxes
    # Filter args from request.form (status)
    filter_args = request.values
    
    # Validate at least one field is selected
    if not selected_fields:
        # In a real app we might flash an error, but for now let's just default to basic fields or error
        # Since user asked to enforce it:
        return "Chyba: Nebyla vybrána žádná pole pro export.", 400

    # Define field definitions (Header Name, Accessor Key)
    # Map field_id -> (Header Label, Dict Key)
    field_definitions = {
        'id': ('ID', 'id'),
        'membership_id': ('Členské číslo', 'membership_id'),
        'first_name': ('Jméno', 'first_name'),
        'last_name': ('Příjmení', 'last_name'),
        'email': ('Email', 'email'),
        'phone': ('Telefon', 'phone'),
        'dob': ('Datum narození', 'dob'),
        'city': ('Město', 'city'),
        'school': ('Škola', 'school'),
        'status': ('Stav', 'status'),
        'application_received': ('Datum přijetí', 'application_received'),
        'created_at': ('Vytvořeno', 'created_at'),
        'interests': ('Zájmy', 'interests'),
        'character': ('Povaha', 'character'),
        'frequency': ('Frekvence', 'frequency'),
        'color': ('Barva', 'color'),
        'source': ('Zdroj', 'source'),
        'source_detail': ('Detail zdroje', 'source_detail'),
        'message': ('Vzkaz', 'message'),
        'newsletter': ('Newsletter', 'newsletter'),
        'guessed_gender': ('Pohlaví (odhad)', 'guessed_gender')
    }
    
    applicants = get_filtered_applicants(filter_args)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Přihlášky"
    
    # Construct Headers based on selection
    headers = []
    keys_to_export = []
    
    for field_id in selected_fields:
        if field_id in field_definitions:
            name, key = field_definitions[field_id]
            headers.append(name)
            keys_to_export.append(key)
            
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
    for app in applicants:
        row_data = []
        for key in keys_to_export:
            val = app.get(key)
            
            # Special formatting for certain keys
            if key == 'dob' and val:
                try:
                    val = datetime.strptime(val.replace('/', '.').strip(), '%d.%m.%Y').date()
                except (ValueError, TypeError):
                    pass
            elif key == 'application_received' and val:
                 try:
                    # Try YYYY-MM-DD HH:MM:SS
                    if '.' in str(val):
                        val = str(val).split('.')[0]
                    val = datetime.strptime(str(val), '%Y-%m-%d %H:%M:%S')
                 except (ValueError, TypeError):
                    pass
            elif key == 'created_at' and val:
                 try:
                    # Created at is already ISO usually? Or matching app received?
                    # Let's assume similar parsing if it's a string
                    if isinstance(val, str):
                        val = datetime.strptime(val, '%Y-%m-%d %H:%M:%S')
                 except (ValueError, TypeError):
                    pass

            row_data.append(val)
            
        ws.append(row_data)

    # Format Date Columns
    # We need to find which indices are date columns dynamically
    date_indices = []
    datetime_indices = []
    
    for idx, key in enumerate(keys_to_export):
        if key == 'dob':
            date_indices.append(idx + 1) # 1-based index
        elif key in ['application_received', 'created_at']:
            datetime_indices.append(idx + 1)
            
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in date_indices:
            cell = row[col_idx-1]
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = 'DD.MM.YYYY'
        
        for col_idx in datetime_indices:
            cell = row[col_idx-1]
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = 'DD.MM.YYYY HH:MM:SS'
             
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, 
        download_name=f'prihlasky_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

# Fetch Logic
@applicants_bp.route('/fetch/preview', methods=['POST'])
@login_required
def fetch_preview():
    """Fetch unread emails preview"""
    import os
    username = os.getenv('EMAIL_USER')
    password = os.getenv('EMAIL_PASS')
    server = os.getenv('IMAP_SERVER', 'imap.gmail.com')
    
    if not username or not password:
         return jsonify({'error': 'Email credentials not configured'}), 500
         
    from src.fetcher import get_unread_emails
    from src.parser import parse_email_body
    
    try:
        raw_emails = get_unread_emails(username, password, server, mark_as_read=False)
        previews = []
        
        session['fetched_emails'] = []
        
        total_count = 0
        new_count = 0
        duplicates_count = 0
        
        conn = get_db_connection()
        # Fetch existing membership IDs (ignoring deleted and empty ones)
        existing_ids = {str(row['membership_id']) for row in conn.execute("SELECT membership_id FROM applicants WHERE deleted = 0 AND membership_id IS NOT NULL AND membership_id != ''").fetchall()}
        conn.close()
        
        for email_uid, body, date in raw_emails:
            total_count += 1
            parsed = parse_email_body(body)
            parsed['email_uid'] = email_uid
            parsed['date'] = date
            
            # Check duplicate by Membership ID
            mem_id = parsed.get('membership_id')
            if mem_id and str(mem_id) in existing_ids:
                duplicates_count += 1
                parsed['is_duplicate'] = True
            else:
                new_count += 1
                parsed['is_duplicate'] = False
                
            previews.append(parsed)
            # Store simple dict in session
            session['fetched_emails'].append({'uid': email_uid, 'body': body, 'date': date})
            
        return jsonify({
            'emails': previews,
            'total': total_count,
            'new': new_count,
            'duplicates': duplicates_count
        })
    except Exception as e:
        logger.error(f"Fetch error: {e}")
        return jsonify({'error': str(e)}), 500

@applicants_bp.route('/fetch/confirm', methods=['POST'])
@login_required
def fetch_confirm():
    """Confirm and save fetched emails"""
    import os
    username = os.getenv('EMAIL_USER')
    password = os.getenv('EMAIL_PASS')
    server = os.getenv('IMAP_SERVER', 'imap.gmail.com')
    
    from src.fetcher import get_unread_emails
    from src.parser import parse_email_body
    
    try:
        # Determine mode
        mode = session.get('mode', 'test')
        should_mark_read = (mode == 'production')
        
        # We process from session cache or re-fetch (marking read this time)
        # Using cache allows us to process exactly what was previewed, but we must mark them read on server eventually.
        # Ideally, we call get_unread_emails(..., mark_as_read=True) to commit.
        
        raw_emails = get_unread_emails(username, password, server, mark_as_read=should_mark_read)
        count = 0
        errors = []
        conn = get_db_connection()
        
        for email_uid, body, date in raw_emails:
            parsed = parse_email_body(body)
            email_addr = parsed.get('email', 'Unknown')
            
            # Check for existing Membership ID (ignore deleted)
            mem_id = parsed.get('membership_id')
            
            if not mem_id:
                errors.append(f"Email {email_addr}: Chybí členské číslo. Nelze vytvořit přihlášku.")
                continue
            
            if mem_id:
                # Check including deleted
                existing = conn.execute('SELECT id, deleted FROM applicants WHERE membership_id = ?', (mem_id,)).fetchone()
                
                if not existing:
                     # Check by email/name unique constraint to avoid crash
                     existing = conn.execute('SELECT id, deleted FROM applicants WHERE email = ? AND first_name = ? AND last_name = ?', 
                                           (email_addr, parsed.get('first_name'), parsed.get('last_name'))).fetchone()
                
                if existing:
                    if existing['deleted']:
                        # Restore
                        conn.execute('UPDATE applicants SET deleted = 0 WHERE id = ?', (existing['id'],))
                        log_action(existing['id'], "Obnoveno z emailu", session.get('user', {}).get('email'), connection=conn)
                        count += 1
                    else:
                        # Log duplicate skip
                        # We log this so it appears in the audit trail that an attempt was made
                        log_action(existing['id'], "Pokus o import z emailu (duplicita)", session.get('user', {}).get('email'), connection=conn)
                    continue
                
            parsed['status'] = 'Nová'
            parsed['application_received'] = datetime.now()
            
            keys = [k for k in parsed.keys() if k != 'full_body' and k in [
                'first_name', 'last_name', 'email', 'phone', 'dob', 'city', 'school', 
                'interests', 'character', 'status', 'newsletter', 'source', 
                'source_detail', 'message', 'color', 'guessed_gender', 'membership_id',
                'application_received'
            ]]
            
            placeholders = ', '.join(['?' for _ in keys])
            cols = ', '.join(keys)
            vals = [parsed[k] for k in keys]
            
            cursor = conn.execute(f'INSERT INTO applicants ({cols}) VALUES ({placeholders})', vals)
            new_id = cursor.lastrowid
            
            log_action(new_id, "Vytvořeno z emailu", session.get('user', {}).get('email'), connection=conn)
            count += 1
            
        conn.commit()
        conn.close()
        
        session.pop('fetched_emails', None)
        return jsonify({'success': True, 'count': count, 'errors': errors})
        
    except Exception as e:
        logger.error(f"Fetch confirm error: {e}")
        return jsonify({'error': str(e)}), 500


